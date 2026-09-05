import io
import openpyxl
from typing import List, Dict, Any

class ExcelProcessor:
    @staticmethod
    def export_to_excel(headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1") -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        for row in rows:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def parse_excel(file_bytes: bytes) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                item = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                data.append(item)
        return data
        
    @staticmethod
    def generate_equipment_template() -> bytes:
        headers = ["设备编码*", "设备名称*", "规格型号*", "位置编码*", "额定电压", "设备参数信息"]
        return ExcelProcessor.export_to_excel(headers, [], sheet_name="设备导入模板")
